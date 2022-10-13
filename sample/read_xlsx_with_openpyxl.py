import openpyxl

dataframe = openpyxl.load_workbook("sales_data.xlsx")

dataframe1 = dataframe.active

# satırdan kolona
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)

# kolondan satıra
for col in dataframe1.iter_cols(1, dataframe1.max_column):
    for row in range(0, dataframe1.max_row):
        print(col[row].value)